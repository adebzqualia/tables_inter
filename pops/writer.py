from __future__ import annotations

from collections import Counter

from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .aggregator import (
    build_derived_formula,
    build_ratio_formula,
    build_value_group_formula,
)
from .config import AppConfig, KPIConfig, SheetConfig
from .lineage import configured_node_id
from .models import (
    ExtractionRecord,
    LineageNode,
    LineagePlan,
    LineageRecord,
    ResolvedRatioTotal,
    ValidationIssue,
)

_HEADER_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=12)
_TOTAL_FONT = Font(bold=True)
_THIN_TOP_BORDER = Border(top=Side(style="thin"))
_HYPERLINK_FONT = Font(underline="single")


def _prepare_generated_sheet(wb: Workbook, name: str, replace_existing: bool) -> Worksheet:
    """Create a generated worksheet, optionally replacing a previous one.

    :param wb: Formula-preserving workbook.
    :param name: Generated worksheet name.
    :param replace_existing: Whether an existing generated sheet may be removed.
    :return: Fresh worksheet.
    """

    if name in wb.sheetnames:
        if not replace_existing:
            raise ValueError(
                f"Generated worksheet {name!r} already exists and replacement is disabled"
            )
        del wb[name]
    return wb.create_sheet(title=name)


def _record_lookup(records: list[ExtractionRecord]) -> dict[tuple[str, int, str, str], ExtractionRecord]:
    """Index configured extraction records."""

    lookup: dict[tuple[str, int, str, str], ExtractionRecord] = {}
    for record in records:
        key = (record.source, record.kpi_index, record.country, record.period)
        if key in lookup:
            raise ValueError(f"Duplicate extraction record detected: {key}")
        lookup[key] = record
    return lookup


def _dependency_record_lookup(records: tuple[LineageRecord, ...]) -> dict[tuple[str, str, str], LineageRecord]:
    """Index automatically discovered dependency records."""

    return {(record.node_id, record.country, record.period): record for record in records}


def _style_header_row(ws: Worksheet, row: int, last_column: int) -> None:
    """Apply simple header formatting."""

    for column in range(1, last_column + 1):
        cell = ws.cell(row=row, column=column)
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _quote_sheet_name(name: str) -> str:
    """Quote an Excel worksheet name for formulas and internal links."""

    return "'" + name.replace("'", "''") + "'"


def _is_percent_format(number_format: str | None) -> bool:
    """Return whether an Excel number format represents a percentage."""

    return bool(number_format and "%" in number_format)


def _rounded_number_format(percent: bool, digits: int) -> str:
    """Build a simple rounded output number format."""

    decimals = "." + ("0" * digits) if digits else ""
    return f"0{decimals}%" if percent else f"#,##0{decimals}"


def _source_formula(
    sheet_name: str,
    coordinate: str,
    round_values: bool,
    round_digits: int,
    percent: bool,
) -> str:
    """Build a guarded Excel formula linking to a source cell."""

    reference = f"{_quote_sheet_name(sheet_name)}!{coordinate}"
    value_expression = reference
    if round_values:
        if percent:
            value_expression = f"ROUND({reference}*100,{round_digits})/100"
        else:
            value_expression = f"ROUND({reference},{round_digits})"
    return f'=IFERROR(IF({reference}="","",{value_expression}),"")'


def _source_hyperlink(sheet_name: str, coordinate: str) -> str:
    """Build an internal hyperlink to a source cell."""

    return f"#{_quote_sheet_name(sheet_name)}!{coordinate}"


def _configured_layout(
    kpis: list[KPIConfig],
    countries: tuple[str, ...],
    groups: dict[str, tuple[str, ...]],
    spacing_rows: int,
) -> tuple[dict[int, dict[str, int]], dict[tuple[int, str], int], dict[int, int]]:
    """Pre-compute rows for configured KPI tables."""

    country_rows: dict[int, dict[str, int]] = {}
    total_rows: dict[tuple[int, str], int] = {}
    title_rows: dict[int, int] = {}
    row = 1
    for kpi in kpis:
        title_rows[kpi.index] = row
        first_country_row = row + 2
        country_rows[kpi.index] = {
            country: first_country_row + offset
            for offset, country in enumerate(countries)
        }
        total_row = first_country_row + len(countries) + 1
        for group_name in groups:
            total_rows[(kpi.index, group_name)] = total_row
            total_row += 1
        row = total_row + spacing_rows
    return country_rows, total_rows, title_rows


def _dependency_layout(
    node_ids: tuple[str, ...],
    countries: tuple[str, ...],
    groups: dict[str, tuple[str, ...]],
    spacing_rows: int,
) -> tuple[dict[str, dict[str, int]], dict[tuple[str, str], int], dict[str, int]]:
    """Pre-compute rows for automatically discovered dependency tables."""

    country_rows: dict[str, dict[str, int]] = {}
    total_rows: dict[tuple[str, str], int] = {}
    title_rows: dict[str, int] = {}
    row = 1
    for node_id in node_ids:
        title_rows[node_id] = row
        first_country_row = row + 2
        country_rows[node_id] = {
            country: first_country_row + offset
            for offset, country in enumerate(countries)
        }
        total_row = first_country_row + len(countries) + 1
        for group_name in groups:
            total_rows[(node_id, group_name)] = total_row
            total_row += 1
        row = total_row + spacing_rows
    return country_rows, total_rows, title_rows


def _majority_percent(records: list[ExtractionRecord], source: str, kpi_index: int) -> bool:
    """Infer whether a configured KPI is percentage-formatted in source cells."""

    flags = [
        _is_percent_format(record.number_format)
        for record in records
        if record.source == source and record.kpi_index == kpi_index and record.coordinate
    ]
    return bool(flags) and sum(flags) >= (len(flags) // 2 + 1)


def _write_aggregation_control_sheet(
    ws: Worksheet,
    config: AppConfig,
    lineage: LineagePlan,
) -> dict[str, str]:
    """Write live SUM/AVERAGE controls and return absolute cell references.

    Ordinary value KPIs receive a dropdown that operational users can change
    directly in Excel. Formula-derived KPIs are informational and non-editable.
    """

    headers = ["Node ID", "Source", "Type", "Subtype", "KPI", "Role", "Aggregation", "Notes"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    controls: dict[str, str] = {}
    editable_cells: list[str] = []
    row = 2

    for source, source_config in config.workbook.sources.items():
        if not source_config.enabled:
            continue
        for kpi in config.kpis_by_source[source]:
            if kpi.aggregation not in {"sum", "average", "ratio"}:
                continue
            node_id = configured_node_id(source, kpi.index)
            is_value = kpi.aggregation in {"sum", "average"}
            aggregation = kpi.aggregation.upper() if is_value else "FORMULA"
            note = (
                "Editable in Excel: choose SUM or AVERAGE for TOP8/TOP9/ALL."
                if is_value
                else "Derived KPI: group totals use an explicit or recursively traced formula."
            )
            ws.append([
                node_id,
                source,
                kpi.type,
                kpi.subtype,
                kpi.name,
                "BUSINESS KPI",
                aggregation,
                note,
            ])
            if is_value:
                cell = ws.cell(row=row, column=7)
                editable_cells.append(cell.coordinate)
                controls[node_id] = (
                    f"{_quote_sheet_name(config.workbook.aggregation_control_sheet)}!$G${row}"
                )
            row += 1

    node_by_id = {node.node_id: node for node in lineage.nodes}
    for node_id in lineage.dependency_order:
        node = node_by_id[node_id]
        aggregation = node.default_aggregation.upper()
        if aggregation in {"SUM", "AVERAGE"}:
            note = "Automatically discovered value KPI. Editable: SUM or AVERAGE."
        elif aggregation == "FORMULA":
            note = "Automatically discovered derived KPI. Formula traced recursively."
        else:
            note = "Formula/ratio semantics could not be reconstructed safely; totals stay blank."
        ws.append([
            node.node_id,
            node.source,
            None,
            None,
            node.name,
            "AUTO DEPENDENCY",
            aggregation,
            note,
        ])
        if aggregation in {"SUM", "AVERAGE"}:
            cell = ws.cell(row=row, column=7)
            editable_cells.append(cell.coordinate)
            controls[node_id] = (
                f"{_quote_sheet_name(config.workbook.aggregation_control_sheet)}!$G${row}"
            )
        row += 1

    validation = DataValidation(type="list", formula1='"SUM,AVERAGE"', allow_blank=False)
    validation.error = "Choose SUM or AVERAGE."
    validation.errorTitle = "Invalid aggregation"
    validation.prompt = "Change this directly in Excel to control group totals."
    validation.promptTitle = "Group aggregation"
    ws.add_data_validation(validation)
    for coordinate in editable_cells:
        validation.add(ws[coordinate])
        ws[coordinate].comment = Comment(
            "Operational control: changing SUM/AVERAGE updates intermediary group totals immediately; no Python rerun is required.",
            "POPS",
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].hidden = True
    widths = {"B": 20, "C": 34, "D": 24, "E": 60, "F": 20, "G": 16, "H": 85}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    return controls


def _resolve_symbolic_formula(
    expression: str,
    group_name: str,
    total_refs: dict[tuple[str, str, str], str],
    lineage: LineagePlan,
) -> tuple[str | None, list[str]]:
    """Replace lineage placeholders with intermediary group-total references."""

    resolved = expression
    dependency_refs: list[str] = []
    for token, (node_id, period) in lineage.placeholder_targets.items():
        if token not in resolved:
            continue
        reference = total_refs.get((node_id, group_name, period))
        if reference is None:
            return None, []
        resolved = resolved.replace(token, reference)
        dependency_refs.append(reference)
    return resolved, dependency_refs


def _write_configured_source_sheet(
    ws: Worksheet,
    source_name: str,
    sheet: SheetConfig,
    config: AppConfig,
    records: list[ExtractionRecord],
    ratio_total_rules: dict[tuple[str, int], ResolvedRatioTotal],
    lineage: LineagePlan,
    aggregation_controls: dict[str, str],
    total_refs: dict[tuple[str, str, str], str],
    layout: tuple[dict[int, dict[str, int]], dict[tuple[int, str], int], dict[int, int]],
) -> None:
    """Write one configured intermediary source sheet."""

    record_by_key = _record_lookup(records)
    configured_kpis = config.kpis_by_source[source_name]
    output_kpis = [
        kpi for kpi in configured_kpis if kpi.aggregation in {"sum", "average", "ratio"}
    ]
    country_rows, total_rows, title_rows = layout
    last_column = 1 + len(sheet.periods)

    for kpi in output_kpis:
        title_row = title_rows[kpi.index]
        title_cell = ws.cell(
            row=title_row,
            column=1,
            value=kpi.display_name(config.workbook.kpi_title_separator),
        )
        title_cell.font = _TITLE_FONT
        if last_column > 1:
            ws.merge_cells(
                start_row=title_row,
                start_column=1,
                end_row=title_row,
                end_column=last_column,
            )

        header_row = title_row + 1
        ws.cell(row=header_row, column=1, value="Countries")
        for offset, period in enumerate(sheet.periods, start=2):
            ws.cell(row=header_row, column=offset, value=period)
        _style_header_row(ws, header_row, last_column)

        for country in config.countries.countries:
            row = country_rows[kpi.index][country]
            ws.cell(row=row, column=1, value=country)
            for offset, period in enumerate(sheet.periods, start=2):
                record = record_by_key[(source_name, kpi.index, country, period)]
                cell = ws.cell(row=row, column=offset)
                if record.coordinate is None:
                    continue

                percent = kpi.aggregation == "ratio" or _is_percent_format(record.number_format)
                digits = (
                    config.workbook.ratio_round_digits
                    if percent
                    else config.workbook.round_digits
                )
                cell.value = _source_formula(
                    record.source_sheet,
                    record.coordinate,
                    config.workbook.round_values,
                    digits,
                    percent,
                )
                if config.workbook.round_values:
                    cell.number_format = _rounded_number_format(percent, digits)
                elif record.number_format:
                    cell.number_format = record.number_format
                if config.workbook.add_source_hyperlinks:
                    cell.hyperlink = _source_hyperlink(record.source_sheet, record.coordinate)
                    cell.font = _HYPERLINK_FONT

        node_id = configured_node_id(source_name, kpi.index)
        explicit_rule = ratio_total_rules.get((source_name, kpi.index))
        percent_total = kpi.aggregation == "ratio" or _majority_percent(records, source_name, kpi.index)

        for group_name, members in config.countries.groups.items():
            total_row = total_rows[(kpi.index, group_name)]
            label_cell = ws.cell(row=total_row, column=1, value=f"TOTAL {group_name}")
            label_cell.font = _TOTAL_FONT
            label_cell.border = _THIN_TOP_BORDER

            for offset, period in enumerate(sheet.periods, start=2):
                cell = ws.cell(row=total_row, column=offset)
                if kpi.aggregation in {"sum", "average"}:
                    control_ref = aggregation_controls[node_id]
                    digits = (
                        config.workbook.ratio_round_digits
                        if percent_total
                        else config.workbook.round_digits
                    )
                    cell.value = build_value_group_formula(
                        offset,
                        country_rows[kpi.index],
                        members,
                        control_ref,
                        config.workbook.round_values,
                        digits,
                    )
                    if config.workbook.round_values:
                        cell.number_format = _rounded_number_format(percent_total, digits)

                elif explicit_rule is not None:
                    numerator_id = configured_node_id(source_name, explicit_rule.numerator_index)
                    denominator_id = configured_node_id(source_name, explicit_rule.denominator_index)
                    numerator_ref = total_refs[(numerator_id, group_name, period)]
                    denominator_ref = total_refs[(denominator_id, group_name, period)]
                    cell.value = build_ratio_formula(
                        numerator_ref,
                        denominator_ref,
                        explicit_rule.multiplier,
                        explicit_rule.percent,
                        config.workbook.round_values,
                        config.workbook.ratio_round_digits,
                    )
                    if explicit_rule.percent:
                        cell.number_format = _rounded_number_format(True, config.workbook.ratio_round_digits)

                else:
                    symbolic = lineage.configured_formulas.get((source_name, kpi.index, period))
                    if symbolic is not None:
                        expression, dependencies = _resolve_symbolic_formula(
                            symbolic, group_name, total_refs, lineage
                        )
                        if expression is not None:
                            cell.value = build_derived_formula(
                                expression,
                                dependencies,
                                percent=True,
                                round_values=config.workbook.round_values,
                                round_digits=config.workbook.ratio_round_digits,
                            )
                            cell.number_format = _rounded_number_format(
                                True, config.workbook.ratio_round_digits
                            )

                cell.font = _TOTAL_FONT
                cell.border = _THIN_TOP_BORDER

    ws.column_dimensions["A"].width = 32
    for column in range(2, last_column + 1):
        period = sheet.periods[column - 2]
        ws.column_dimensions[get_column_letter(column)].width = max(12, len(period) + 2)


def _write_dependency_sheet(
    ws: Worksheet,
    config: AppConfig,
    lineage: LineagePlan,
    aggregation_controls: dict[str, str],
    total_refs: dict[tuple[str, str, str], str],
    layout: tuple[dict[str, dict[str, int]], dict[tuple[str, str], int], dict[str, int]],
) -> None:
    """Write only the recursively discovered supporting KPI tables."""

    if not lineage.dependency_order:
        ws["A1"] = "No additional KPI dependencies were required."
        return

    nodes = {node.node_id: node for node in lineage.nodes}
    records = _dependency_record_lookup(lineage.records)
    country_rows, total_rows, title_rows = layout
    periods = config.workbook.sources[config.workbook.kpi_source_name].periods
    last_column = 1 + len(periods)

    for node_id in lineage.dependency_order:
        node = nodes[node_id]
        title_row = title_rows[node_id]
        title = ws.cell(row=title_row, column=1, value=node.display_name)
        title.font = _TITLE_FONT
        if last_column > 1:
            ws.merge_cells(
                start_row=title_row,
                start_column=1,
                end_row=title_row,
                end_column=last_column,
            )

        header_row = title_row + 1
        ws.cell(row=header_row, column=1, value="Countries")
        for offset, period in enumerate(periods, start=2):
            ws.cell(row=header_row, column=offset, value=period)
        _style_header_row(ws, header_row, last_column)

        for country in config.countries.countries:
            row = country_rows[node_id][country]
            ws.cell(row=row, column=1, value=country)
            for offset, period in enumerate(periods, start=2):
                record = records.get((node_id, country, period))
                if record is None or record.coordinate is None:
                    continue
                cell = ws.cell(row=row, column=offset)
                percent = node.percent or _is_percent_format(record.number_format)
                digits = config.workbook.ratio_round_digits if percent else config.workbook.round_digits
                cell.value = _source_formula(
                    record.source_sheet,
                    record.coordinate,
                    config.workbook.round_values,
                    digits,
                    percent,
                )
                if config.workbook.round_values:
                    cell.number_format = _rounded_number_format(percent, digits)
                elif record.number_format:
                    cell.number_format = record.number_format
                if config.workbook.add_source_hyperlinks:
                    cell.hyperlink = _source_hyperlink(record.source_sheet, record.coordinate)
                    cell.font = _HYPERLINK_FONT

        for group_name, members in config.countries.groups.items():
            row = total_rows[(node_id, group_name)]
            label = ws.cell(row=row, column=1, value=f"TOTAL {group_name}")
            label.font = _TOTAL_FONT
            label.border = _THIN_TOP_BORDER

            for offset, period in enumerate(periods, start=2):
                cell = ws.cell(row=row, column=offset)
                symbolic = lineage.formulas.get((node_id, period))
                if symbolic is not None:
                    expression, dependencies = _resolve_symbolic_formula(
                        symbolic, group_name, total_refs, lineage
                    )
                    if expression is not None:
                        cell.value = build_derived_formula(
                            expression,
                            dependencies,
                            percent=node.percent,
                            round_values=config.workbook.round_values,
                            round_digits=(
                                config.workbook.ratio_round_digits
                                if node.percent
                                else config.workbook.round_digits
                            ),
                        )
                elif node.default_aggregation in {"sum", "average"}:
                    control_ref = aggregation_controls[node_id]
                    cell.value = build_value_group_formula(
                        offset,
                        country_rows[node_id],
                        members,
                        control_ref,
                        config.workbook.round_values,
                        config.workbook.round_digits,
                    )

                if node.percent:
                    cell.number_format = _rounded_number_format(
                        True, config.workbook.ratio_round_digits
                    )
                elif config.workbook.round_values:
                    cell.number_format = _rounded_number_format(
                        False, config.workbook.round_digits
                    )
                cell.font = _TOTAL_FONT
                cell.border = _THIN_TOP_BORDER

    ws.column_dimensions["A"].width = 40
    for column in range(2, last_column + 1):
        ws.column_dimensions[get_column_letter(column)].width = 14


def _write_validation_sheet(ws: Worksheet, issues: list[ValidationIssue]) -> None:
    """Write validation diagnostics to a dedicated worksheet."""

    headers = ["Country", "Source Sheet", "KPI", "Period", "Issue", "Cell/Details"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for issue in issues:
        ws.append([
            issue.country,
            issue.source_sheet,
            issue.kpi,
            issue.period,
            issue.issue,
            issue.details,
        ])

    widths = {"A": 16, "B": 28, "C": 65, "D": 16, "E": 40, "F": 105}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_generated_sheets(
    wb: Workbook,
    config: AppConfig,
    records: list[ExtractionRecord],
    ratio_total_rules: dict[tuple[str, int], ResolvedRatioTotal],
    lineage: LineagePlan,
    issues: list[ValidationIssue],
) -> None:
    """Write business tables, recursive dependencies, live controls and validation."""

    replace_existing = config.workbook.replace_existing_generated_sheets

    configured_layouts = {}
    total_refs: dict[tuple[str, str, str], str] = {}
    for source_name, sheet in config.workbook.sources.items():
        if not sheet.enabled:
            continue
        output_kpis = [
            kpi
            for kpi in config.kpis_by_source[source_name]
            if kpi.aggregation in {"sum", "average", "ratio"}
        ]
        layout = _configured_layout(
            output_kpis,
            config.countries.countries,
            config.countries.groups,
            config.workbook.table_spacing_rows,
        )
        configured_layouts[source_name] = layout
        _country_rows, total_rows, _title_rows = layout
        if sheet.output_sheet is None:
            raise ValueError(f"Enabled source {source_name!r} has no output sheet")
        for kpi in output_kpis:
            node_id = configured_node_id(source_name, kpi.index)
            for group_name in config.countries.groups:
                row = total_rows[(kpi.index, group_name)]
                for column, period in enumerate(sheet.periods, start=2):
                    total_refs[(node_id, group_name, period)] = (
                        f"{_quote_sheet_name(sheet.output_sheet)}!{get_column_letter(column)}{row}"
                    )

    dependency_layout = _dependency_layout(
        lineage.dependency_order,
        config.countries.countries,
        config.countries.groups,
        config.workbook.table_spacing_rows,
    )
    _dep_country_rows, dep_total_rows, _dep_title_rows = dependency_layout
    dependency_periods = config.workbook.sources[config.workbook.kpi_source_name].periods
    for node_id in lineage.dependency_order:
        for group_name in config.countries.groups:
            row = dep_total_rows[(node_id, group_name)]
            for column, period in enumerate(dependency_periods, start=2):
                total_refs[(node_id, group_name, period)] = (
                    f"{_quote_sheet_name(config.workbook.dependency_sheet)}!"
                    f"{get_column_letter(column)}{row}"
                )

    control_ws = _prepare_generated_sheet(
        wb,
        config.workbook.aggregation_control_sheet,
        replace_existing,
    )
    aggregation_controls = _write_aggregation_control_sheet(control_ws, config, lineage)

    for source_name, sheet in config.workbook.sources.items():
        if not sheet.enabled:
            continue
        ws = _prepare_generated_sheet(wb, sheet.output_sheet, replace_existing)
        _write_configured_source_sheet(
            ws=ws,
            source_name=source_name,
            sheet=sheet,
            config=config,
            records=records,
            ratio_total_rules=ratio_total_rules,
            lineage=lineage,
            aggregation_controls=aggregation_controls,
            total_refs=total_refs,
            layout=configured_layouts[source_name],
        )

    dependency_ws = _prepare_generated_sheet(
        wb,
        config.workbook.dependency_sheet,
        replace_existing,
    )
    _write_dependency_sheet(
        dependency_ws,
        config,
        lineage,
        aggregation_controls,
        total_refs,
        dependency_layout,
    )

    validation_ws = _prepare_generated_sheet(
        wb,
        config.workbook.validation_sheet,
        replace_existing,
    )
    _write_validation_sheet(validation_ws, issues)

    calculation = getattr(wb, "calculation", None)
    if calculation is not None:
        calculation.calcMode = "auto"
        calculation.fullCalcOnLoad = True
        calculation.forceFullCalc = True
