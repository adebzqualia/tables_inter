from __future__ import annotations

import hashlib
import re
from collections import Counter, defaultdict, deque
from dataclasses import dataclass
from typing import Iterable

from openpyxl.formula import Tokenizer
from openpyxl.utils import get_column_letter, range_boundaries

from .config import AppConfig, KPIConfig, normalize_label
from .models import (
    ExtractionRecord,
    LineageNode,
    LineagePlan,
    LineageRecord,
    ValidationIssue,
)

_CELL_RE = re.compile(r"^\$?[A-Z]{1,3}\$?\d+$", re.IGNORECASE)
_AGGREGATE_RANGE_FUNCTIONS = {"SUM", "AVERAGE", "MIN", "MAX", "COUNT", "COUNTA"}


def configured_node_id(source: str, kpi_index: int) -> str:
    """Build the stable node identifier for a configured KPI.

    :param source: Logical source sheet.
    :param kpi_index: Zero-based configured KPI index.
    :return: Stable node identifier.
    """

    return f"cfg:{source}:{kpi_index}"


@dataclass
class _NodeState:
    """Mutable state for one automatically discovered KPI dependency."""

    node_id: str
    source: str
    name: str
    occurrence: int
    display_name: str
    percent: bool = False
    default_aggregation: str = "sum"
    saw_formula: bool = False
    formula_periods: int = 0


@dataclass(frozen=True)
class _KpiCatalog:
    """Resolved rectangular catalogue information for one physical KPI sheet."""

    sheet_name: str
    kpi_column: int
    header_row: int
    period_columns: dict[str, int]
    period_by_column: dict[int, str]
    rows_by_name: dict[str, tuple[int, ...]]


class _LineageResolver:
    """Recursively trace formula dependencies and build symbolic total formulas."""

    def __init__(self, formulas_wb, values_wb, config: AppConfig, records: list[ExtractionRecord]):
        self.formulas_wb = formulas_wb
        self.values_wb = values_wb
        self.config = config
        self.records = records
        self.issues: list[ValidationIssue] = []

        self._configured_by_cell: dict[tuple[str, str], tuple[str, str, KPIConfig]] = {}
        self._configured_nodes: dict[str, tuple[str, KPIConfig]] = {}
        for record in records:
            if record.coordinate is None:
                continue
            node_id = configured_node_id(record.source, record.kpi_index)
            kpi = config.kpis_by_source[record.source][record.kpi_index]
            self._configured_by_cell[(record.source_sheet, self._coord(record.coordinate))] = (
                node_id,
                record.period,
                kpi,
            )
            self._configured_nodes[node_id] = (record.source, kpi)

        self._catalogs: dict[str, _KpiCatalog | None] = {}
        self._node_states: dict[str, _NodeState] = {}
        self._node_records: dict[tuple[str, str, str], LineageRecord] = {}
        self._node_queue: deque[str] = deque()
        self._node_queued: set[str] = set()
        self._node_analyzed: set[str] = set()

        self._formulas: dict[tuple[str, str], str] = {}
        self._configured_formulas: dict[tuple[str, int, str], str] = {}
        self._placeholder_targets: dict[str, tuple[str, str]] = {}
        self._placeholder_by_target: dict[tuple[str, str], str] = {}
        self._formula_failures: Counter[str] = Counter()

        self._sheet_identity: dict[str, tuple[str, str]] = {}
        for country in config.countries.countries:
            for source in config.workbook.sources:
                physical = config.workbook.sheet_name_template.format(
                    country=country,
                    sheet=source,
                )
                self._sheet_identity[physical] = (country, source)

    @staticmethod
    def _coord(value: str) -> str:
        """Normalize an A1 coordinate for dictionary lookups."""

        return value.replace("$", "").upper()

    @staticmethod
    def _unquote_sheet(value: str) -> str:
        """Unquote an Excel sheet token."""

        value = value.strip()
        if value.startswith("'") and value.endswith("'"):
            return value[1:-1].replace("''", "'")
        return value

    def _resolve_sheet_reference(self, country: str, current_sheet: str, raw_sheet: str | None) -> str | None:
        """Resolve an Excel formula sheet reference inside the consolidated workbook."""

        if raw_sheet is None:
            return current_sheet

        sheet_name = self._unquote_sheet(raw_sheet)
        if "[" in sheet_name or "]" in sheet_name:
            return None
        if sheet_name in self.formulas_wb.sheetnames:
            return sheet_name

        candidate = self.config.workbook.sheet_name_template.format(
            country=country,
            sheet=sheet_name,
        )
        if candidate in self.formulas_wb.sheetnames:
            return candidate
        return None

    def _placeholder(self, node_id: str, period: str) -> str:
        """Return a stable symbolic placeholder for one KPI total reference."""

        target = (node_id, period)
        existing = self._placeholder_by_target.get(target)
        if existing is not None:
            return existing

        token = f"__POPS_REF_{len(self._placeholder_targets) + 1:05d}__"
        self._placeholder_targets[token] = target
        self._placeholder_by_target[target] = token
        return token

    def _build_kpi_catalog(self, sheet_name: str) -> _KpiCatalog | None:
        """Locate the KPI column and configured period columns in a KPI sheet."""

        if sheet_name in self._catalogs:
            return self._catalogs[sheet_name]
        if sheet_name not in self.formulas_wb.sheetnames:
            self._catalogs[sheet_name] = None
            return None

        ws = self.formulas_wb[sheet_name]
        target = normalize_label(self.config.workbook.kpi_name_header)
        header_candidates = []
        period_norms = {
            normalize_label(period): period
            for period in self.config.workbook.sources[self.config.workbook.kpi_source_name].periods
        }

        period_cells: dict[str, list[tuple[int, int]]] = defaultdict(list)
        for row in ws.iter_rows():
            for cell in row:
                normalized = normalize_label(cell.value)
                if normalized == target:
                    header_candidates.append(cell)
                if normalized in period_norms:
                    period_cells[period_norms[normalized]].append((cell.row, cell.column))

        if not header_candidates:
            self._catalogs[sheet_name] = None
            return None

        def score(cell) -> tuple[int, int, int]:
            close_periods = sum(
                1
                for refs in period_cells.values()
                if any(abs(row - cell.row) <= 2 for row, _column in refs)
            )
            same_row = sum(
                1
                for refs in period_cells.values()
                if any(row == cell.row for row, _column in refs)
            )
            return same_row, close_periods, -cell.row

        header = max(header_candidates, key=score)
        period_columns: dict[str, int] = {}
        for period, refs in period_cells.items():
            ranked = sorted(
                refs,
                key=lambda item: (
                    abs(item[0] - header.row),
                    0 if item[0] <= header.row + 2 else 1,
                    item[0],
                    item[1],
                ),
            )
            if ranked:
                period_columns[period] = ranked[0][1]

        rows_by_name: dict[str, list[int]] = defaultdict(list)
        for row in range(header.row + 1, ws.max_row + 1):
            label = ws.cell(row=row, column=header.column).value
            normalized = normalize_label(label)
            if normalized:
                rows_by_name[normalized].append(row)

        catalog = _KpiCatalog(
            sheet_name=sheet_name,
            kpi_column=header.column,
            header_row=header.row,
            period_columns=period_columns,
            period_by_column={column: period for period, column in period_columns.items()},
            rows_by_name={name: tuple(rows) for name, rows in rows_by_name.items()},
        )
        self._catalogs[sheet_name] = catalog
        return catalog

    def _auto_descriptor(self, country: str, sheet_name: str, coordinate: str) -> tuple[str, str, int, str] | None:
        """Map a KPI-sheet cell to a semantic auto dependency node."""

        identity = self._sheet_identity.get(sheet_name)
        if identity is None or identity[0] != country or identity[1] != self.config.workbook.kpi_source_name:
            return None

        catalog = self._build_kpi_catalog(sheet_name)
        if catalog is None:
            return None

        cell = self.formulas_wb[sheet_name][self._coord(coordinate)]
        period = catalog.period_by_column.get(cell.column)
        if period is None:
            return None

        label = self.formulas_wb[sheet_name].cell(row=cell.row, column=catalog.kpi_column).value
        normalized = normalize_label(label)
        if not normalized:
            return None

        rows = catalog.rows_by_name.get(normalized, ())
        if cell.row not in rows:
            return None
        occurrence = rows.index(cell.row) + 1
        digest = hashlib.sha1(normalized.encode("utf-8")).hexdigest()[:12]
        node_id = f"auto:{self.config.workbook.kpi_source_name}:{occurrence}:{digest}"
        return node_id, str(label).strip(), occurrence, period

    def _create_auto_node(self, node_id: str, name: str, occurrence: int) -> None:
        """Create and populate a recursively discovered KPI dependency node."""

        if node_id in self._node_states:
            return

        source = self.config.workbook.kpi_source_name
        periods = self.config.workbook.sources[source].periods
        state = _NodeState(
            node_id=node_id,
            source=source,
            name=name,
            occurrence=occurrence,
            display_name=f"{source} | {name}" + (f" [#{occurrence}]" if occurrence > 1 else ""),
            default_aggregation=self.config.workbook.default_dependency_aggregation,
        )
        self._node_states[node_id] = state

        percent_flags: list[bool] = []
        for country in self.config.countries.countries:
            physical = self.config.workbook.sheet_name_template.format(country=country, sheet=source)
            catalog = self._build_kpi_catalog(physical)
            if catalog is None:
                for period in periods:
                    self._node_records[(node_id, country, period)] = LineageRecord(
                        node_id=node_id,
                        country=country,
                        period=period,
                        source_sheet=physical,
                        coordinate=None,
                    )
                continue

            rows = catalog.rows_by_name.get(normalize_label(name), ())
            row = rows[occurrence - 1] if len(rows) >= occurrence else None
            for period in periods:
                column = catalog.period_columns.get(period)
                if row is None or column is None:
                    self._node_records[(node_id, country, period)] = LineageRecord(
                        node_id=node_id,
                        country=country,
                        period=period,
                        source_sheet=physical,
                        coordinate=None,
                    )
                    continue

                formula_cell = self.formulas_wb[physical].cell(row=row, column=column)
                value_cell = self.values_wb[physical].cell(row=row, column=column)
                value = value_cell.value
                numeric_value = value if isinstance(value, (int, float)) and not isinstance(value, bool) else None
                formula = formula_cell.value if formula_cell.data_type == "f" else None
                number_format = formula_cell.number_format or "General"
                percent_flags.append("%" in number_format)
                self._node_records[(node_id, country, period)] = LineageRecord(
                    node_id=node_id,
                    country=country,
                    period=period,
                    source_sheet=physical,
                    coordinate=formula_cell.coordinate,
                    number_format=number_format,
                    formula=formula if isinstance(formula, str) else None,
                    value=numeric_value,
                )

        state.percent = bool(percent_flags) and sum(percent_flags) >= (len(percent_flags) // 2 + 1)
        if node_id not in self._node_queued:
            self._node_queue.append(node_id)
            self._node_queued.add(node_id)

    def _compile_reference(
        self,
        country: str,
        sheet_name: str,
        coordinate: str,
        stack: tuple[tuple[str, str], ...],
        depth: int,
    ) -> str | None:
        """Compile one referenced source cell to a semantic total expression."""

        coordinate = self._coord(coordinate)
        configured = self._configured_by_cell.get((sheet_name, coordinate))
        if configured is not None:
            node_id, period, _kpi = configured
            return self._placeholder(node_id, period)

        descriptor = self._auto_descriptor(country, sheet_name, coordinate)
        if descriptor is not None:
            node_id, name, occurrence, period = descriptor
            self._create_auto_node(node_id, name, occurrence)
            return self._placeholder(node_id, period)

        if depth >= self.config.workbook.formula_max_depth:
            self._formula_failures["maximum recursion depth reached"] += 1
            return None

        cell = self.formulas_wb[sheet_name][coordinate]
        if cell.data_type == "f" and isinstance(cell.value, str):
            return self._compile_formula_cell(
                country=country,
                sheet_name=sheet_name,
                coordinate=coordinate,
                stack=stack,
                depth=depth + 1,
            )

        value = cell.value
        if value is None:
            self._formula_failures["blank helper cell"] += 1
            return None
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        if isinstance(value, (int, float)):
            return f"{value:g}"
        if isinstance(value, str):
            escaped = value.replace('"', '""')
            return f'"{escaped}"'

        self._formula_failures["unsupported helper-cell value type"] += 1
        return None

    def _parse_range_token(
        self,
        country: str,
        current_sheet: str,
        token_value: str,
        aggregate_function: str | None,
        stack: tuple[tuple[str, str], ...],
        depth: int,
    ) -> str | None:
        """Compile one Tokenizer RANGE operand, including simple aggregate ranges."""

        raw = token_value.strip()
        if "!" in raw:
            raw_sheet, raw_address = raw.rsplit("!", 1)
        else:
            raw_sheet, raw_address = None, raw

        sheet_name = self._resolve_sheet_reference(country, current_sheet, raw_sheet)
        if sheet_name is None:
            self._formula_failures["unresolved or external worksheet reference"] += 1
            return None

        address = raw_address.replace("$", "").upper()
        if ":" not in address:
            if not _CELL_RE.match(address):
                self._formula_failures["named/structured reference is unsupported"] += 1
                return None
            return self._compile_reference(country, sheet_name, address, stack, depth)

        if aggregate_function not in _AGGREGATE_RANGE_FUNCTIONS:
            self._formula_failures["cell range outside a supported aggregate function"] += 1
            return None

        try:
            min_col, min_row, max_col, max_row = range_boundaries(address)
        except ValueError:
            self._formula_failures["invalid range reference"] += 1
            return None

        if (max_col - min_col + 1) * (max_row - min_row + 1) > 200:
            self._formula_failures["formula range is too large to trace safely"] += 1
            return None

        compiled: list[str] = []
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                coordinate = f"{get_column_letter(column)}{row}"
                expression = self._compile_reference(country, sheet_name, coordinate, stack, depth)
                if expression is None:
                    return None
                compiled.append(f"({expression})")
        return ",".join(compiled)

    def _compile_formula_cell(
        self,
        country: str,
        sheet_name: str,
        coordinate: str,
        stack: tuple[tuple[str, str], ...] = (),
        depth: int = 0,
    ) -> str | None:
        """Recursively compile an Excel formula into semantic KPI placeholders."""

        coordinate = self._coord(coordinate)
        key = (sheet_name, coordinate)
        if key in stack:
            self._formula_failures["circular formula dependency"] += 1
            return None
        if depth > self.config.workbook.formula_max_depth:
            self._formula_failures["maximum recursion depth reached"] += 1
            return None

        cell = self.formulas_wb[sheet_name][coordinate]
        if cell.data_type != "f" or not isinstance(cell.value, str):
            return self._compile_reference(country, sheet_name, coordinate, stack, depth)

        try:
            tokens = Tokenizer(cell.value).items
        except Exception:
            self._formula_failures["formula tokenizer failed"] += 1
            return None

        result: list[str] = []
        function_stack: list[str] = []
        new_stack = stack + (key,)

        for token in tokens:
            if token.type == "FUNC" and token.subtype == "OPEN":
                function_name = token.value[:-1].strip().upper()
                function_stack.append(function_name)
                result.append(token.value)
                continue
            if token.type == "FUNC" and token.subtype == "CLOSE":
                if function_stack:
                    function_stack.pop()
                result.append(token.value)
                continue
            if token.type == "OPERAND" and token.subtype == "RANGE":
                expression = self._parse_range_token(
                    country=country,
                    current_sheet=sheet_name,
                    token_value=token.value,
                    aggregate_function=function_stack[-1] if function_stack else None,
                    stack=new_stack,
                    depth=depth,
                )
                if expression is None:
                    return None
                result.append(f"({expression})")
                continue
            result.append(token.value)

        return "".join(result).strip()

    @staticmethod
    def _consensus(expressions: dict[str, set[str]], minimum: int) -> tuple[str | None, tuple[str, ...], str | None]:
        """Resolve one cross-country formula consensus."""

        conflicts = [country for country, values in expressions.items() if len(values) > 1]
        if conflicts:
            return None, (), f"different formulas within {', '.join(conflicts)}"

        one_per_country = {
            country: next(iter(values))
            for country, values in expressions.items()
            if len(values) == 1
        }
        distinct = set(one_per_country.values())
        if len(distinct) != 1:
            if len(distinct) > 1:
                return None, (), "different semantic formulas across countries"
            return None, (), "no formula could be traced"
        if len(one_per_country) < minimum:
            return None, (), f"only {len(one_per_country)} country/countries confirmed the formula"

        formula = next(iter(distinct))
        evidence = tuple(one_per_country)
        return formula, evidence, None

    def _compile_configured_ratio(self, source: str, kpi: KPIConfig, period: str) -> None:
        """Compile one configured ratio period across countries."""

        expressions: dict[str, set[str]] = defaultdict(set)
        for record in self.records:
            if (
                record.source != source
                or record.kpi_index != kpi.index
                or record.period != period
                or record.coordinate is None
            ):
                continue
            cell = self.formulas_wb[record.source_sheet][record.coordinate]
            if cell.data_type != "f":
                continue
            expression = self._compile_formula_cell(
                country=record.country,
                sheet_name=record.source_sheet,
                coordinate=record.coordinate,
            )
            if expression:
                expressions[record.country].add(expression)

        minimum = 1 if len(self.config.countries.countries) == 1 else 2
        formula, evidence, reason = self._consensus(expressions, minimum)
        if formula is not None:
            self._configured_formulas[(source, kpi.index, period)] = formula
            self.issues.append(
                ValidationIssue(
                    country=None,
                    source_sheet=source,
                    kpi=kpi.display_name(self.config.workbook.kpi_title_separator),
                    period=period,
                    issue="RECURSIVE_FORMULA_TOTAL_INFERRED",
                    details=(
                        f"Group-total formula traced recursively and confirmed in "
                        f"{len(evidence)}/{len(self.config.countries.countries)} countries."
                    ),
                )
            )
        else:
            self.issues.append(
                ValidationIssue(
                    country=None,
                    source_sheet=source,
                    kpi=kpi.display_name(self.config.workbook.kpi_title_separator),
                    period=period,
                    issue="RECURSIVE_FORMULA_TOTAL_UNRESOLVED",
                    details=f"{reason}. This period's group totals remain blank.",
                )
            )

    def _analyze_auto_node(self, node_id: str) -> None:
        """Determine whether an auto dependency is a value or formula-derived KPI."""

        if node_id in self._node_analyzed:
            return
        state = self._node_states[node_id]
        periods = self.config.workbook.sources[state.source].periods
        minimum = 1 if len(self.config.countries.countries) == 1 else 2
        total_formula_cells = 0
        successful_periods = 0

        for period in periods:
            expressions: dict[str, set[str]] = defaultdict(set)
            formula_cells = 0
            for country in self.config.countries.countries:
                record = self._node_records.get((node_id, country, period))
                if record is None or record.coordinate is None or not record.formula:
                    continue
                formula_cells += 1
                expression = self._compile_formula_cell(
                    country=country,
                    sheet_name=record.source_sheet,
                    coordinate=record.coordinate,
                )
                if expression:
                    expressions[country].add(expression)

            total_formula_cells += formula_cells
            if formula_cells == 0:
                continue

            formula, evidence, reason = self._consensus(expressions, minimum)
            if formula is not None:
                self._formulas[(node_id, period)] = formula
                successful_periods += 1
            else:
                self.issues.append(
                    ValidationIssue(
                        country=None,
                        source_sheet=state.source,
                        kpi=state.display_name,
                        period=period,
                        issue="DEPENDENCY_FORMULA_UNRESOLVED",
                        details=f"{reason}. Dependency total remains blank for this period.",
                    )
                )

        state.saw_formula = total_formula_cells > 0
        state.formula_periods = successful_periods
        if successful_periods:
            state.default_aggregation = "formula"
        elif state.saw_formula:
            state.default_aggregation = "unresolved"
        elif state.percent:
            state.default_aggregation = "unresolved"
            self.issues.append(
                ValidationIssue(
                    country=None,
                    source_sheet=state.source,
                    kpi=state.display_name,
                    period=None,
                    issue="DEPENDENCY_RATIO_WITHOUT_FORMULA",
                    details=(
                        "The discovered KPI is percentage-formatted but no traceable formula "
                        "was found. Its country values are shown, but group totals are left blank."
                    ),
                )
            )
        else:
            state.default_aggregation = self.config.workbook.default_dependency_aggregation

        self._node_analyzed.add(node_id)

    def _dependency_order(self) -> tuple[str, ...]:
        """Topologically order automatic dependencies with leaves first."""

        auto_ids = set(self._node_states)
        dependencies: dict[str, set[str]] = {node_id: set() for node_id in auto_ids}
        reverse: dict[str, set[str]] = {node_id: set() for node_id in auto_ids}

        for (node_id, _period), formula in self._formulas.items():
            for token, (target_id, _target_period) in self._placeholder_targets.items():
                if token in formula and target_id in auto_ids and target_id != node_id:
                    dependencies[node_id].add(target_id)
                    reverse[target_id].add(node_id)

        indegree = {node_id: len(dependencies[node_id]) for node_id in auto_ids}
        ready = deque(sorted(node_id for node_id, degree in indegree.items() if degree == 0))
        ordered: list[str] = []
        while ready:
            node_id = ready.popleft()
            ordered.append(node_id)
            for dependent in sorted(reverse[node_id]):
                indegree[dependent] -= 1
                if indegree[dependent] == 0:
                    ready.append(dependent)

        if len(ordered) != len(auto_ids):
            remaining = sorted(auto_ids - set(ordered))
            ordered.extend(remaining)
            self.issues.append(
                ValidationIssue(
                    country=None,
                    source_sheet=self.config.workbook.dependency_sheet,
                    kpi=None,
                    period=None,
                    issue="DEPENDENCY_GRAPH_CYCLE",
                    details=(
                        "A cycle exists in the discovered dependency graph. Cyclic nodes are "
                        "still displayed, but their group formulas may remain unresolved."
                    ),
                )
            )
        return tuple(ordered)

    def build(self) -> LineagePlan:
        """Build the recursive lineage plan for configured ratio KPIs."""

        if self.config.workbook.recursive_formula_totals:
            for source, source_config in self.config.workbook.sources.items():
                if not source_config.enabled:
                    continue
                for kpi in self.config.kpis_by_source[source]:
                    if kpi.aggregation != "ratio" or kpi.ratio_total is not None:
                        continue
                    for period in source_config.periods:
                        self._compile_configured_ratio(source, kpi, period)

            while self._node_queue:
                node_id = self._node_queue.popleft()
                self._analyze_auto_node(node_id)

        nodes = tuple(
            LineageNode(
                node_id=state.node_id,
                source=state.source,
                name=state.name,
                occurrence=state.occurrence,
                display_name=state.display_name,
                configured=False,
                configured_index=None,
                default_aggregation=state.default_aggregation,
                percent=state.percent,
            )
            for state in self._node_states.values()
        )

        if self._formula_failures:
            details = ", ".join(
                f"{reason}: {count}"
                for reason, count in self._formula_failures.most_common(8)
            )
            self.issues.append(
                ValidationIssue(
                    country=None,
                    source_sheet=None,
                    kpi=None,
                    period=None,
                    issue="FORMULA_LINEAGE_LIMITATIONS",
                    details=f"Some formula branches could not be traced safely ({details}).",
                )
            )

        return LineagePlan(
            nodes=nodes,
            records=tuple(self._node_records.values()),
            formulas=dict(self._formulas),
            configured_formulas=dict(self._configured_formulas),
            placeholder_targets=dict(self._placeholder_targets),
            dependency_order=self._dependency_order(),
        )


def build_lineage_plan(
    formulas_wb,
    values_wb,
    config: AppConfig,
    records: list[ExtractionRecord],
) -> tuple[LineagePlan, list[ValidationIssue]]:
    """Trace derived-KPI formulas recursively and discover required KPI tables.

    The resolver follows cell references across the consolidated workbook. A
    referenced KPI-sheet cell is identified through the configured ``KPI`` name
    column, materialized as an automatic dependency table, and recursively traced
    if that cell is itself formula-derived.

    :param formulas_wb: Workbook loaded with formulas preserved.
    :param values_wb: Workbook loaded with cached formula values.
    :param config: Complete application configuration.
    :param records: Configured KPI extraction records.
    :return: Lineage plan and validation diagnostics.
    """

    resolver = _LineageResolver(formulas_wb, values_wb, config, records)
    plan = resolver.build()
    return plan, resolver.issues
