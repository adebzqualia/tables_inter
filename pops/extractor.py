from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from decimal import Decimal
from typing import Iterable

from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from .config import AppConfig, KPIConfig, SheetConfig, normalize_label
from .models import ExtractionRecord, ValidationIssue

_EXCEL_ERROR_VALUES = {
    "#NULL!",
    "#DIV/0!",
    "#VALUE!",
    "#REF!",
    "#NAME?",
    "#NUM!",
    "#N/A",
    "#GETTING_DATA",
}


@dataclass(frozen=True)
class CellRef:
    """Coordinate and raw label value for an indexed cell.

    :param row: One-based Excel row.
    :param column: One-based Excel column.
    :param coordinate: Excel coordinate such as ``J37``.
    :param value: Raw indexed value.
    """

    row: int
    column: int
    coordinate: str
    value: object


@dataclass(frozen=True)
class ResolvedLabels:
    """Resolved KPI and period labels for one worksheet.

    :param periods: Period label to resolved cell, or ``None``.
    :param kpis: KPI configuration index to resolved cell, or ``None``.
    :param issues: Matching diagnostics.
    """

    periods: dict[str, CellRef | None]
    kpis: dict[int, CellRef | None]
    issues: tuple[ValidationIssue, ...]


def _physical_sheet_name(template: str, country: str, source: str) -> str:
    """Construct a physical source worksheet name.

    :param template: Configured sheet-name template.
    :param country: Country/entity name.
    :param source: Logical source-sheet name.
    :return: Physical worksheet name.
    """

    return template.format(country=country, sheet=source)


def build_label_index(
    values_ws: Worksheet,
    formulas_ws: Worksheet,
) -> dict[str, list[CellRef]]:
    """Scan a worksheet once and index normalized cell labels.

    Formula cells use the ``data_only`` workbook's cached result. Non-formula
    cells use the formula-preserving workbook directly.

    :param values_ws: Worksheet from the ``data_only=True`` workbook.
    :param formulas_ws: Worksheet from the formula-preserving workbook.
    :return: Mapping from normalized label to all matching coordinates.
    """

    index: dict[str, list[CellRef]] = defaultdict(list)

    for row in formulas_ws.iter_rows():
        for formula_cell in row:
            raw_value = (
                values_ws[formula_cell.coordinate].value
                if formula_cell.data_type == "f"
                else formula_cell.value
            )
            normalized = normalize_label(raw_value)
            if not normalized:
                continue

            index[normalized].append(
                CellRef(
                    row=formula_cell.row,
                    column=formula_cell.column,
                    coordinate=formula_cell.coordinate,
                    value=raw_value,
                )
            )

    for refs in index.values():
        refs.sort(key=lambda ref: (ref.row, ref.column))

    return dict(index)


def _support_by_row(candidates: dict[str, list[CellRef]]) -> Counter[int]:
    """Count how many distinct configured labels occur on each row.

    :param candidates: Candidate coordinates keyed by label.
    :return: Row support counts.
    """

    support: Counter[int] = Counter()
    for refs in candidates.values():
        for row in {ref.row for ref in refs}:
            support[row] += 1
    return support


def _coordinates(refs: Iterable[CellRef]) -> str:
    """Format candidate coordinates for diagnostics.

    :param refs: Candidate cells.
    :return: Comma-separated coordinates.
    """

    return ", ".join(ref.coordinate for ref in refs)


def _resolve_periods(
    country: str,
    sheet_name: str,
    periods: tuple[str, ...],
    label_index: dict[str, list[CellRef]],
    kpis: tuple[KPIConfig, ...],
) -> tuple[dict[str, CellRef | None], list[ValidationIssue]]:
    """Resolve configured period labels using exact normalized matching.

    :param country: Country/entity name.
    :param sheet_name: Physical worksheet name.
    :param periods: Ordered configured periods.
    :param label_index: One-pass worksheet label index.
    :param kpis: Configured KPIs, used only to infer table-body context.
    :return: Resolved period cells and diagnostics.
    """

    candidates = {
        period: list(label_index.get(normalize_label(period), []))
        for period in periods
    }
    row_support = _support_by_row(candidates)

    all_kpi_rows = [
        ref.row
        for kpi in kpis
        for ref in label_index.get(normalize_label(kpi.name), [])
    ]
    body_top = min(all_kpi_rows) if all_kpi_rows else None

    resolved: dict[str, CellRef | None] = {}
    issues: list[ValidationIssue] = []

    for period in periods:
        refs = candidates[period]
        if not refs:
            resolved[period] = None
            issues.append(
                ValidationIssue(
                    country=country,
                    source_sheet=sheet_name,
                    kpi=None,
                    period=period,
                    issue="PERIOD_NOT_FOUND",
                    details="Configured period label was not found.",
                )
            )
            continue

        if len(refs) == 1:
            resolved[period] = refs[0]
            continue

        contextual = refs
        if body_top is not None:
            above_body = [ref for ref in refs if ref.row < body_top]
            if above_body:
                contextual = above_body

        max_support = max(row_support[ref.row] for ref in contextual)
        strongest = [
            ref for ref in contextual if row_support[ref.row] == max_support
        ]
        chosen = strongest[0] if len(strongest) == 1 and max_support > 1 else None

        if chosen is None:
            resolved[period] = None
            issues.append(
                ValidationIssue(
                    country=country,
                    source_sheet=sheet_name,
                    kpi=None,
                    period=period,
                    issue="DUPLICATE_PERIOD_MATCH",
                    details=f"Ambiguous exact matches at: {_coordinates(refs)}",
                )
            )
        else:
            resolved[period] = chosen
            issues.append(
                ValidationIssue(
                    country=country,
                    source_sheet=sheet_name,
                    kpi=None,
                    period=period,
                    issue="DUPLICATE_PERIOD_RESOLVED",
                    details=(
                        f"Exact matches at: {_coordinates(refs)}. "
                        f"Resolved by header-row context to {chosen.coordinate}."
                    ),
                )
            )

    return resolved, issues


def _resolve_kpis(
    country: str,
    sheet_name: str,
    kpis: tuple[KPIConfig, ...],
    label_index: dict[str, list[CellRef]],
    resolved_periods: dict[str, CellRef | None],
    title_separator: str,
) -> tuple[dict[int, CellRef | None], list[ValidationIssue]]:
    """Resolve KPI occurrences positionally in worksheet order.

    Same-name KPI definitions are grouped by normalized name. If the source
    worksheet contains the same number of contextual matches, they are assigned
    in configuration order to top-to-bottom worksheet occurrences. A count
    mismatch leaves the whole same-name group unresolved to avoid positional
    shifting after a missing/extra occurrence.

    :param country: Country/entity name.
    :param sheet_name: Physical worksheet name.
    :param kpis: Ordered configured KPI occurrences.
    :param label_index: One-pass worksheet label index.
    :param resolved_periods: Resolved period headers for table context.
    :param title_separator: Separator used in display labels.
    :return: Resolved KPI cells and diagnostics.
    """

    period_refs = [ref for ref in resolved_periods.values() if ref is not None]
    header_bottom = max((ref.row for ref in period_refs), default=None)
    leftmost_period_column = min((ref.column for ref in period_refs), default=None)

    grouped: dict[str, list[KPIConfig]] = defaultdict(list)
    for kpi in kpis:
        grouped[normalize_label(kpi.name)].append(kpi)

    resolved: dict[int, CellRef | None] = {kpi.index: None for kpi in kpis}
    issues: list[ValidationIssue] = []

    for normalized_name, configured in grouped.items():
        refs = list(label_index.get(normalized_name, []))
        contextual = [
            ref
            for ref in refs
            if (header_bottom is None or ref.row > header_bottom)
            and (
                leftmost_period_column is None
                or ref.column < leftmost_period_column
            )
        ]
        if contextual:
            refs = contextual
        refs.sort(key=lambda ref: (ref.row, ref.column))

        expected = len(configured)
        found = len(refs)
        representative = configured[0]

        if found == 0:
            for kpi in configured:
                issues.append(
                    ValidationIssue(
                        country=country,
                        source_sheet=sheet_name,
                        kpi=kpi.display_name(title_separator),
                        period=None,
                        issue="KPI_NOT_FOUND",
                        details="Configured KPI label was not found.",
                    )
                )
            continue

        if found != expected:
            display_names = "; ".join(
                kpi.display_name(title_separator) for kpi in configured
            )
            issues.append(
                ValidationIssue(
                    country=country,
                    source_sheet=sheet_name,
                    kpi=representative.name,
                    period=None,
                    issue="KPI_OCCURRENCE_COUNT_MISMATCH",
                    details=(
                        f"Expected {expected} occurrence(s) of {representative.name!r} "
                        f"from configuration, found {found} at: {_coordinates(refs)}. "
                        f"Configured occurrences: {display_names}. No positional "
                        "assignment was made for this name."
                    ),
                )
            )
            continue

        for kpi, ref in zip(configured, refs):
            resolved[kpi.index] = ref

    return resolved, issues


def resolve_labels(
    country: str,
    sheet_name: str,
    sheet: SheetConfig,
    kpis: tuple[KPIConfig, ...],
    label_index: dict[str, list[CellRef]],
    title_separator: str,
) -> ResolvedLabels:
    """Resolve all configured periods and KPI occurrences for one worksheet.

    :param country: Country/entity name.
    :param sheet_name: Physical worksheet name.
    :param sheet: Source-sheet configuration.
    :param kpis: KPI configuration for the source.
    :param label_index: One-pass worksheet label index.
    :param title_separator: Separator used in KPI display labels.
    :return: Resolved labels and diagnostics.
    """

    periods, period_issues = _resolve_periods(
        country=country,
        sheet_name=sheet_name,
        periods=sheet.periods,
        label_index=label_index,
        kpis=kpis,
    )
    resolved_kpis, kpi_issues = _resolve_kpis(
        country=country,
        sheet_name=sheet_name,
        kpis=kpis,
        label_index=label_index,
        resolved_periods=periods,
        title_separator=title_separator,
    )

    return ResolvedLabels(
        periods=periods,
        kpis=resolved_kpis,
        issues=tuple(period_issues + kpi_issues),
    )


def _merged_anchor(ws: Worksheet, row: int, column: int) -> Cell | MergedCell:
    """Return the top-left anchor when a coordinate lies in a merged range.

    :param ws: Worksheet.
    :param row: One-based row.
    :param column: One-based column.
    :return: Real cell or merged-range anchor.
    """

    cell = ws.cell(row=row, column=column)
    if not isinstance(cell, MergedCell):
        return cell

    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell


def _is_excel_error(cell: Cell | MergedCell, value: object) -> bool:
    """Return whether a cell/value represents an Excel error.

    :param cell: Resolved worksheet cell.
    :param value: Cell value.
    :return: ``True`` for Excel error values.
    """

    if getattr(cell, "data_type", None) == "e":
        return True
    return isinstance(value, str) and value.strip().upper() in _EXCEL_ERROR_VALUES


def _extract_numeric_value(
    country: str,
    sheet_name: str,
    kpi_display: str,
    period: str,
    row: int,
    column: int,
    values_ws: Worksheet,
    formulas_ws: Worksheet,
) -> tuple[int | float | None, str, str, ValidationIssue | None]:
    """Inspect one additive KPI value while retaining its source coordinate.

    :param country: Country/entity name.
    :param sheet_name: Physical worksheet name.
    :param kpi_display: KPI display label.
    :param period: Period label.
    :param row: Resolved KPI row.
    :param column: Resolved period column.
    :param values_ws: Worksheet containing cached formula results.
    :param formulas_ws: Formula-preserving worksheet.
    :return: Numeric cached value, source coordinate, source number format, and optional diagnostic.
    """

    formula_cell = _merged_anchor(formulas_ws, row, column)
    value_cell = _merged_anchor(values_ws, row, column)
    coordinate = formula_cell.coordinate
    number_format = getattr(formula_cell, "number_format", "General") or "General"
    raw_value = value_cell.value

    if _is_excel_error(value_cell, raw_value) or _is_excel_error(
        formula_cell, formula_cell.value
    ):
        return None, coordinate, number_format, ValidationIssue(
            country=country,
            source_sheet=sheet_name,
            kpi=kpi_display,
            period=period,
            issue="EXCEL_ERROR_VALUE",
            details=f"Excel error at {coordinate}: {raw_value!r}",
        )

    if getattr(formula_cell, "data_type", None) == "f" and raw_value is None:
        return None, coordinate, number_format, ValidationIssue(
            country=country,
            source_sheet=sheet_name,
            kpi=kpi_display,
            period=period,
            issue="FORMULA_WITHOUT_CACHED_VALUE",
            details=(
                f"Formula at {coordinate} has no cached calculated value. "
                "The generated intermediary formula will still link to it, but "
                "the input workbook should be recalculated/saved in Excel."
            ),
        )

    if raw_value is None:
        return None, coordinate, number_format, ValidationIssue(
            country=country,
            source_sheet=sheet_name,
            kpi=kpi_display,
            period=period,
            issue="BLANK_VALUE",
            details=f"Resolved source cell {coordinate} is blank.",
        )

    if isinstance(raw_value, bool) or not isinstance(raw_value, (int, float, Decimal)):
        return None, coordinate, number_format, ValidationIssue(
            country=country,
            source_sheet=sheet_name,
            kpi=kpi_display,
            period=period,
            issue="NON_NUMERIC_VALUE",
            details=f"Expected a numeric value at {coordinate}, found {raw_value!r}.",
        )

    if isinstance(raw_value, Decimal):
        raw_value = float(raw_value)
    return raw_value, coordinate, number_format, None


def _empty_records_for_sheet(
    country: str,
    sheet_name: str,
    source: str,
    periods: tuple[str, ...],
    kpis: tuple[KPIConfig, ...],
    title_separator: str,
) -> list[ExtractionRecord]:
    """Create unresolved records for all additive KPI intersections.

    :param country: Country/entity name.
    :param sheet_name: Expected physical worksheet name.
    :param source: Logical source-sheet name.
    :param periods: Configured periods.
    :param kpis: Configured KPIs.
    :param title_separator: Separator used in display labels.
    :return: Missing extraction records for generated ``sum`` and ``ratio`` KPIs.
    """

    return [
        ExtractionRecord(
            country=country,
            source=source,
            source_sheet=sheet_name,
            kpi_index=kpi.index,
            kpi=kpi.name,
            kpi_display=kpi.display_name(title_separator),
            period=period,
            value=None,
            coordinate=None,
            number_format=None,
        )
        for kpi in kpis
        if kpi.aggregation in {"sum", "ratio"}
        for period in periods
    ]


def extract_workbook(
    values_wb,
    formulas_wb,
    config: AppConfig,
) -> tuple[list[ExtractionRecord], list[ValidationIssue]]:
    """Extract all enabled source sheets from the consolidated workbook.

    Relevant worksheets are scanned once. Repeated KPI names are mapped by
    configuration order to worksheet order. Additive and ratio KPIs produce output
    records; ``skip`` KPIs are resolved but not output.

    :param values_wb: Workbook loaded with ``data_only=True``.
    :param formulas_wb: Workbook loaded with ``data_only=False``.
    :param config: Complete application configuration.
    :return: Extraction records and validation issues.
    """

    records: list[ExtractionRecord] = []
    issues: list[ValidationIssue] = []
    separator = config.workbook.kpi_title_separator

    for source_name, sheet in config.workbook.sources.items():
        if not sheet.enabled:
            continue

        kpis = config.kpis_by_source[source_name]
        output_kpis = tuple(kpi for kpi in kpis if kpi.aggregation in {"sum", "ratio"})

        for country in config.countries.countries:
            sheet_name = _physical_sheet_name(
                config.workbook.sheet_name_template,
                country,
                source_name,
            )

            if sheet_name not in formulas_wb.sheetnames or sheet_name not in values_wb.sheetnames:
                issues.append(
                    ValidationIssue(
                        country=country,
                        source_sheet=sheet_name,
                        kpi=None,
                        period=None,
                        issue="SOURCE_SHEET_MISSING",
                        details="Expected source worksheet does not exist.",
                    )
                )
                records.extend(
                    _empty_records_for_sheet(
                        country=country,
                        sheet_name=sheet_name,
                        source=source_name,
                        periods=sheet.periods,
                        kpis=kpis,
                        title_separator=separator,
                    )
                )
                continue

            formulas_ws = formulas_wb[sheet_name]
            values_ws = values_wb[sheet_name]
            label_index = build_label_index(values_ws, formulas_ws)
            resolved = resolve_labels(
                country=country,
                sheet_name=sheet_name,
                sheet=sheet,
                kpis=kpis,
                label_index=label_index,
                title_separator=separator,
            )
            issues.extend(resolved.issues)

            for kpi in output_kpis:
                kpi_ref = resolved.kpis[kpi.index]
                kpi_display = kpi.display_name(separator)

                for period in sheet.periods:
                    period_ref = resolved.periods[period]
                    if kpi_ref is None or period_ref is None:
                        records.append(
                            ExtractionRecord(
                                country=country,
                                source=source_name,
                                source_sheet=sheet_name,
                                kpi_index=kpi.index,
                                kpi=kpi.name,
                                kpi_display=kpi_display,
                                period=period,
                                value=None,
                                coordinate=None,
                                number_format=None,
                            )
                        )
                        continue

                    value, coordinate, number_format, issue = _extract_numeric_value(
                        country=country,
                        sheet_name=sheet_name,
                        kpi_display=kpi_display,
                        period=period,
                        row=kpi_ref.row,
                        column=period_ref.column,
                        values_ws=values_ws,
                        formulas_ws=formulas_ws,
                    )
                    records.append(
                        ExtractionRecord(
                            country=country,
                            source=source_name,
                            source_sheet=sheet_name,
                            kpi_index=kpi.index,
                            kpi=kpi.name,
                            kpi_display=kpi_display,
                            period=period,
                            value=value,
                            coordinate=coordinate,
                            number_format=number_format,
                        )
                    )
                    if issue is not None:
                        issues.append(issue)

    return records, issues
