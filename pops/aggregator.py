from __future__ import annotations

from openpyxl.utils import get_column_letter


def build_sum_formula(
    column: int,
    row_by_country: dict[str, int],
    members: tuple[str, ...],
    round_values: bool = False,
    round_digits: int = 0,
) -> str | None:
    """Build an Excel ``SUM`` formula for a configured country group.

    :param column: One-based intermediary worksheet column.
    :param row_by_country: Mapping from country to intermediary row number.
    :param members: Ordered configured group members.
    :param round_values: Whether to round the generated total.
    :param round_digits: Decimal digits used when rounding.
    :return: Excel formula, or ``None`` for an empty group.
    """

    references = [
        f"{get_column_letter(column)}{row_by_country[country]}"
        for country in members
    ]
    if not references:
        return None

    expression = f"SUM({','.join(references)})"
    if round_values:
        expression = f"ROUND({expression},{round_digits})"
    return f'=IF(COUNT({",".join(references)})=0,"",{expression})'


def build_value_group_formula(
    column: int,
    row_by_country: dict[str, int],
    members: tuple[str, ...],
    aggregation_cell_ref: str,
    round_values: bool = False,
    round_digits: int = 0,
) -> str | None:
    """Build a live SUM/AVERAGE total controlled by an Excel config cell.

    Blanks remain blanks: if no numeric country observation exists, the group
    total is blank rather than zero.

    :param column: One-based intermediary worksheet column.
    :param row_by_country: Mapping from country to intermediary row number.
    :param members: Ordered configured group members.
    :param aggregation_cell_ref: Absolute reference to the SUM/AVERAGE dropdown.
    :param round_values: Whether to round the generated group value.
    :param round_digits: Decimal digits used when rounding.
    :return: Excel formula, or ``None`` for an empty group.
    """

    references = [
        f"{get_column_letter(column)}{row_by_country[country]}"
        for country in members
    ]
    if not references:
        return None

    args = ",".join(references)
    sum_expr = f"SUM({args})"
    avg_expr = f"AVERAGE({args})"
    if round_values:
        sum_expr = f"ROUND({sum_expr},{round_digits})"
        avg_expr = f"ROUND({avg_expr},{round_digits})"

    return (
        f'=IF(COUNT({args})=0,"",'
        f'IF(UPPER({aggregation_cell_ref})="AVERAGE",{avg_expr},{sum_expr}))'
    )


def build_ratio_formula(
    numerator_ref: str,
    denominator_ref: str,
    multiplier: float = 1.0,
    percent: bool = False,
    round_values: bool = False,
    round_digits: int = 0,
) -> str:
    """Build a guarded Excel formula for a simple group-level ratio.

    :param numerator_ref: Intermediary total-cell reference for the numerator.
    :param denominator_ref: Intermediary total-cell reference for the denominator.
    :param multiplier: Optional scale applied after division.
    :param percent: Whether the result is stored as an Excel percentage decimal.
    :param round_values: Whether to round the ratio result.
    :param round_digits: Decimal digits used when rounding.
    :return: Excel formula.
    """

    expression = f"({numerator_ref}/{denominator_ref})"
    if multiplier != 1.0:
        expression = f"({expression}*{multiplier:g})"

    if round_values:
        if percent:
            expression = f"(ROUND(({expression})*100,{round_digits})/100)"
        else:
            expression = f"ROUND({expression},{round_digits})"

    return (
        f'=IF(OR({numerator_ref}="",{denominator_ref}=""),"",'
        f'IFERROR({expression},""))'
    )


def build_derived_formula(
    expression: str,
    dependency_refs: list[str],
    percent: bool,
    round_values: bool,
    round_digits: int,
) -> str:
    """Wrap a recursively traced expression as a guarded Excel total formula.

    :param expression: Excel expression without a leading ``=``.
    :param dependency_refs: Intermediary total references required by expression.
    :param percent: Whether the result is an Excel percentage decimal.
    :param round_values: Whether to round the result.
    :param round_digits: Decimal digits used when rounding.
    :return: Complete Excel formula.
    """

    result = expression
    if round_values:
        if percent:
            result = f"ROUND(({result})*100,{round_digits})/100"
        else:
            result = f"ROUND(({result}),{round_digits})"

    unique_refs = list(dict.fromkeys(dependency_refs))
    if unique_refs:
        missing = ",".join(f'{reference}=""' for reference in unique_refs)
        return f'=IF(OR({missing}),"",IFERROR({result},""))'
    return f'=IFERROR({result},"")'
