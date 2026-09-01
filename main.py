from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from pops.config import load_app_config
from pops.extractor import extract_workbook
from pops.models import ValidationIssue
from pops.lineage import build_lineage_plan
from pops.ratio_inference import resolve_ratio_total_rules
from pops.writer import write_generated_sheets

CONFIG_DIR = Path(__file__).resolve().parent / "config"


def _configuration_warnings(config) -> list[ValidationIssue]:
    """Create non-fatal diagnostics for incomplete optional business groups.

    :param config: Complete application configuration.
    :return: Configuration warnings suitable for the validation sheet.
    """

    issues = [
        ValidationIssue(
            country=None,
            source_sheet=None,
            kpi=None,
            period=None,
            issue="EMPTY_COUNTRY_GROUP",
            details=(
                f"Country group {group_name!r} has no members. "
                "Its generated TOTAL row will remain blank until configured."
            ),
        )
        for group_name, members in config.countries.groups.items()
        if not members
    ]

    return issues


def main() -> None:
    """Generate formula-linked intermediary KPI tables in a workbook copy.

    :raises FileNotFoundError: If the configured input workbook does not exist.
    """

    config = load_app_config(CONFIG_DIR)
    input_path = config.runtime.input_workbook
    output_path = config.runtime.output_workbook

    if not input_path.is_file():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    keep_vba = input_path.suffix.lower() == ".xlsm"

    formulas_wb = load_workbook(
        input_path,
        data_only=False,
        keep_vba=keep_vba,
        keep_links=True,
    )
    values_wb = load_workbook(
        input_path,
        data_only=True,
        keep_vba=keep_vba,
        keep_links=True,
    )

    records, issues = extract_workbook(
        values_wb=values_wb,
        formulas_wb=formulas_wb,
        config=config,
    )
    ratio_total_rules, ratio_issues = resolve_ratio_total_rules(
        formulas_wb=formulas_wb,
        config=config,
        records=records,
    )
    lineage, lineage_issues = build_lineage_plan(
        formulas_wb=formulas_wb,
        values_wb=values_wb,
        config=config,
        records=records,
    )
    issues = _configuration_warnings(config) + issues + ratio_issues + lineage_issues

    write_generated_sheets(
        wb=formulas_wb,
        config=config,
        records=records,
        ratio_total_rules=ratio_total_rules,
        lineage=lineage,
        issues=issues,
    )
    formulas_wb.save(output_path)

    ratio_count = sum(
        1
        for source_name, source in config.workbook.sources.items()
        if source.enabled
        for kpi in config.kpis_by_source[source_name]
        if kpi.aggregation == "ratio"
    )
    skip_count = sum(
        1
        for source_name, source in config.workbook.sources.items()
        if source.enabled
        for kpi in config.kpis_by_source[source_name]
        if kpi.aggregation == "skip"
    )

    print(f"Created: {output_path}")
    print(f"Validation issues: {len(issues)}")
    print("Intermediary country values: Excel links to resolved source cells")
    configured_ratio_totals = sum(
        1
        for source_name, source in config.workbook.sources.items()
        if source.enabled
        for kpi in config.kpis_by_source[source_name]
        if kpi.aggregation == "ratio" and kpi.ratio_total is not None
    )
    inferred_ratio_totals = sum(
        1 for rule in ratio_total_rules.values() if rule.origin == "inferred"
    )
    recursive_ratio_periods = len(lineage.configured_formulas)
    dependency_count = len(lineage.nodes)
    print("Value country-group totals: live Excel SUM/AVERAGE formulas")
    print(f"Ratio KPI tables generated: {ratio_count}")
    print(f"Ratio KPIs with configured group-total rules: {configured_ratio_totals}")
    print(f"Ratio KPIs with inferred simple group-total rules: {inferred_ratio_totals}")
    print(f"Recursively traced ratio period formulas: {recursive_ratio_periods}")
    print(f"Automatically discovered dependency KPIs: {dependency_count}")
    print(f"Live aggregation controls: {config.workbook.aggregation_control_sheet}")
    print(f"Explicitly skipped KPIs: {skip_count}")


if __name__ == "__main__":
    main()
