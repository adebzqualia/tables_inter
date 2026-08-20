from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .aggregator import build_sum_formula
from .config import AppConfig, SheetConfig
from .models import ExtractionRecord, ValidationIssue

_HEADER_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=12)
_TOTAL_FONT = Font(bold=True)
_THIN_TOP_BORDER = Border(top=Side(style="thin"))
_HYPERLINK_FONT = Font(underline="single")


def _prepare_generated_sheet(
    wb: Workbook,
    name: str,
    replace_existing: bool,
) -> Worksheet:
    """Create a generated worksheet, optionally replacing a previous one.

    :param wb: Formula-preserving workbook.
    :param name: Generated worksheet name.
    :param replace_existing: Whether an existing generated sheet may be removed.
    :return: Fresh worksheet.
    :raises ValueError: If the sheet already exists and replacement is disabled.
    """

    if name in wb.sheetnames:
        if not replace_existing:
            raise ValueError(
                f"Generated worksheet {name!r} already exists and replacement is disabled"
            )
        del wb[name]
    return wb.create_sheet(title=name)


def _record_lookup(
    records: list[ExtractionRecord],
) -> dict[tuple[str, int, str, str], ExtractionRecord]:
    """Index extraction records using KPI configuration position.

    :param records: Extraction records.
    :return: Lookup keyed by source, KPI index, country, and period.
    :raises ValueError: If a logical observation occurs more than once.
    """

    lookup: dict[tuple[str, int, str, str], ExtractionRecord] = {}
    for record in records:
        key = (record.source, record.kpi_index, record.country, record.period)
        if key in lookup:
            raise ValueError(f"Duplicate extraction record detected: {key}")
        lookup[key] = record
    return lookup


def _style_header_row(ws: Worksheet, row: int, last_column: int) -> None:
    """Apply simple header formatting.

    :param ws: Output worksheet.
    :param row: Header row number.
    :param last_column: Last table column.
    """

    for column in range(1, last_column + 1):
        cell = ws.cell(row=row, column=column)
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _quote_sheet_name(name: str) -> str:
    """Quote an Excel worksheet name for formulas and internal links.

    :param name: Worksheet name.
    :return: Safely quoted worksheet name.
    """

    return "'" + name.replace("'", "''") + "'"


def _source_formula(sheet_name: str, coordinate: str) -> str:
    """Build a guarded Excel formula linking to a source cell.

    A plain direct reference displays an empty Excel source cell as zero. The
    guard preserves genuine blanks and source errors as blank intermediary cells,
    while a real numeric zero remains zero. Non-numeric source text remains
    visible and is ignored by Excel ``SUM`` formulas.

    :param sheet_name: Physical source worksheet name.
    :param coordinate: Source cell coordinate.
    :return: Excel formula.
    """

    reference = f"{_quote_sheet_name(sheet_name)}!{coordinate}"
    return f'=IFERROR(IF({reference}="","",{reference}),"")'


def _source_hyperlink(sheet_name: str, coordinate: str) -> str:
    """Build an internal hyperlink to a source cell.

    :param sheet_name: Physical source worksheet name.
    :param coordinate: Source cell coordinate.
    :return: Internal Excel hyperlink target.
    """

    return f"#{_quote_sheet_name(sheet_name)}!{coordinate}"


def _write_source_sheet(
    ws: Worksheet,
    source_name: str,
    sheet: SheetConfig,
    config: AppConfig,
    records: list[ExtractionRecord],
) -> None:
    """Write all additive KPI tables using native Excel formulas.

    Country cells reference the resolved source cells. Country-group totals are
    native Excel ``SUM`` formulas over the intermediary country rows.

    :param ws: Generated intermediary worksheet.
    :param source_name: Logical source-sheet name.
    :param sheet: Source-sheet configuration.
    :param config: Complete application configuration.
    :param records: Extraction records.
    """

    record_by_key = _record_lookup(records)
    sum_kpis = [
        kpi for kpi in config.kpis_by_source[source_name] if kpi.aggregation == "sum"
    ]

    last_column = 1 + len(sheet.periods)
    row = 1

    for kpi in sum_kpis:
        title = kpi.display_name(config.workbook.kpi_title_separator)
        title_cell = ws.cell(row=row, column=1, value=title)
        title_cell.font = _TITLE_FONT
        if last_column > 1:
            ws.merge_cells(
                start_row=row,
                start_column=1,
                end_row=row,
                end_column=last_column,
            )

        header_row = row + 1
        ws.cell(row=header_row, column=1, value="Countries")
        for offset, period in enumerate(sheet.periods, start=2):
            ws.cell(row=header_row, column=offset, value=period)
        _style_header_row(ws, header_row, last_column)

        data_row = header_row + 1
        row_by_country: dict[str, int] = {}

        for country in config.countries.countries:
            row_by_country[country] = data_row
            ws.cell(row=data_row, column=1, value=country)

            for offset, period in enumerate(sheet.periods, start=2):
                record = record_by_key[(source_name, kpi.index, country, period)]
                cell = ws.cell(row=data_row, column=offset)

                if record.coordinate is not None:
                    cell.value = _source_formula(record.source_sheet, record.coordinate)
                    if config.workbook.add_source_hyperlinks:
                        cell.hyperlink = _source_hyperlink(
                            record.source_sheet,
                            record.coordinate,
                        )
                        cell.font = _HYPERLINK_FONT
                else:
                    cell.value = None

            data_row += 1

        total_row = data_row + 1
        for group_name, members in config.countries.groups.items():
            label_cell = ws.cell(
                row=total_row,
                column=1,
                value=f"TOTAL {group_name}",
            )
            label_cell.font = _TOTAL_FONT
            label_cell.border = _THIN_TOP_BORDER

            for offset, _period in enumerate(sheet.periods, start=2):
                cell = ws.cell(row=total_row, column=offset)
                cell.value = build_sum_formula(offset, row_by_country, members)
                cell.font = _TOTAL_FONT
                cell.border = _THIN_TOP_BORDER

            total_row += 1

        row = total_row + config.workbook.table_spacing_rows

    ws.column_dimensions["A"].width = 32
    for column in range(2, last_column + 1):
        period = sheet.periods[column - 2]
        ws.column_dimensions[get_column_letter(column)].width = max(12, len(period) + 2)


def _write_validation_sheet(
    ws: Worksheet,
    issues: list[ValidationIssue],
) -> None:
    """Write validation diagnostics to a dedicated worksheet.

    :param ws: Validation worksheet.
    :param issues: Ordered validation issues.
    """

    headers = [
        "Country",
        "Source Sheet",
        "KPI",
        "Period",
        "Issue",
        "Cell/Details",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for issue in issues:
        ws.append(
            [
                issue.country,
                issue.source_sheet,
                issue.kpi,
                issue.period,
                issue.issue,
                issue.details,
            ]
        )

    widths = {
        "A": 16,
        "B": 28,
        "C": 65,
        "D": 16,
        "E": 36,
        "F": 100,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_generated_sheets(
    wb: Workbook,
    config: AppConfig,
    records: list[ExtractionRecord],
    issues: list[ValidationIssue],
) -> None:
    """Add formula-linked intermediary and validation sheets to the workbook.

    Existing source worksheets are not recreated or rewritten. Only configured
    generated sheets are created/replaced.

    :param wb: Formula-preserving workbook to modify and save.
    :param config: Complete application configuration.
    :param records: Extraction records containing resolved source coordinates.
    :param issues: Validation issues.
    """

    replace_existing = config.workbook.replace_existing_generated_sheets

    for source_name, sheet in config.workbook.sources.items():
        if not sheet.enabled:
            continue
        if sheet.output_sheet is None:
            raise ValueError(f"Enabled source {source_name!r} has no output sheet")

        ws = _prepare_generated_sheet(
            wb,
            name=sheet.output_sheet,
            replace_existing=replace_existing,
        )
        _write_source_sheet(
            ws=ws,
            source_name=source_name,
            sheet=sheet,
            config=config,
            records=records,
        )

    validation_ws = _prepare_generated_sheet(
        wb,
        name=config.workbook.validation_sheet,
        replace_existing=replace_existing,
    )
    _write_validation_sheet(validation_ws, issues)

    calculation = getattr(wb, "calculation", None)
    if calculation is not None:
        calculation.calcMode = "auto"
        calculation.fullCalcOnLoad = True
        calculation.forceFullCalc = True
